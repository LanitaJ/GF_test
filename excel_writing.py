import scraping
import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font


OUT_XLSX_FILENAME = 'kwork_parsing.xlsx'

def dump_to_xlsx():
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = 1
    names = scraping.scrap_name()
    fprices = scraping.scrap_first_price()
    sprices = scraping.scrap_second_price()
    details = scraping.scrap_detail()
    data_times_sugges = scraping.scrap_time_sugges()
    #Writing all data to excel file
    for name, fprice, sprice, detail, time, sugges in names, fprices, sprices, details, data_times_sugges.keys(), data_times_sugges.values():
        sheet.cell(row=0, column=1, value='Название')
        sheet.cell(row=0, column=2, value='Желаемая цена')
        sheet.cell(row=0, column=3, value='Допустимая цена')
        sheet.cell(row=0, column=4, value='Время на выполнение')
        sheet.cell(row=0, column=5, value='Количество предложений')
        sheet.cell(row=0, column=6, value='Описание')
        sheet.cell(row=num, column=1, value=name)
        sheet.cell(row=num, column=2, value=fprice)
        sheet.cell(row=num, column=3, value=sprice)
        sheet.cell(row=num, column=4, value=time)
        sheet.cell(row=num, column=5, value=sugges)
        sheet.cell(row=num, column=6, value=detail)
        num += 1
    book.save(OUT_XLSX_FILENAME)

if __name__ == '__main__':
    dump_to_xlsx()

