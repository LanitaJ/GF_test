import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font
import re


OUT_XLSX_FILENAME = 'profi_ru.xlsx'

def soup_connection(link):
    res = requests.get(link, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'})
    html = BeautifulSoup(res.text, 'html.parser')
    return html

def scrap_service(link, tag, class_):
    soup = soup_connection(link)
    service = soup.find(tag, class_=class_)
    return service.text

def scrap_category(link, tag, class_):
    categories = []
    soup = soup_connection(link)
    for headers in soup.find_all('h1', class_='_19nVNdl _3uSrgN0'):
        # for category in headers.find_all(tag, class_= class_):
        res_category = ''.join([i for i in headers.text if not i.isdigit()])
        res_category = res_category.replace('\xa0', '')
        categories.append(res_category)
    del categories[0], categories[0]
    return categories

def scrap_services_menu(link):
    data = dict()
    soup = soup_connection(link)
    for service in soup.find_all('a', class_='services-catalog__menu-item'):
        res_service = ''.join([i for i in service.text if not i.isdigit()])
        res_service = res_service.replace('\xa0', '')
        href = str((service['href']))
        href = f'https://profi.ru{href}'
        data[href] = res_service
    return data

def parsing_info(link):
    subcategories = []
    hrefs = []
    categories = scrap_category(link, 'a', 'services-catalog__item')
    categories_count = 0
    soup = soup_connection(link)
    for services in soup.find_all('div', class_='services-catalog__cols'):
        for serviceh in services:
            for headers in serviceh.find_all('h2', class_='_2RgNS0h _3uSrgN0'):
                for subcategory in headers.find_all('a', class_='services-catalog__item'):
                    subcategories.append(subcategory.text)
                    hrefs.append(f"https://profi.ru{str((subcategory['href']))}")
        dump_to_xlsx(OUT_XLSX_FILENAME, categories[categories_count], subcategories, link, scrap_prices(hrefs))
        categories_count += 1
        subcategories = []
        hrefs = []


def dump_to_xlsx(filename, category, subcategories, link, prices):
    book = openpyxl.load_workbook(filename=filename)
    sheet: worksheet = book.worksheets[0]
    prices_count = 0

    sheet.cell(row=1, column=1, value='Категории')
    sheet['A1'].font = Font(bold=True)
    max = sheet.max_row
    sheet.cell(row=max + 1, column=1, value=scrap_services_menu(link)[link])
    sheet[f'A{max + 1}'].font = Font(bold=True)
    sheet.cell(row=max + 2, column=1, value=category)
    sheet[f'A{max + 2}'].font = Font(bold=True, italic=True)
    for row, subcategory in enumerate(subcategories, max + 3):
        sheet.cell(row=row, column=1, value=subcategory)
        sheet.cell(row=row, column=2, value=prices[prices_count])
        prices_count += 1

    book.save(filename)

def scrap_prices(hrefs):
    prices = []
    for href in hrefs:
        reg = re.compile(r'\([^)]*\)')
        href = re.sub(reg, '', str(href))
        soup = soup_connection(href)
        price = None
        if price is None:
            price = soup.find('h3', class_='ui_2MYaG ui_2SXfw ui_RF7aD ui_9c6iR _2xaOGZk')
        if price is None:
            price = soup.find('span', class_='ui_2MYaG ui_2SXfw ui_3nAGW ui_3Y-uv _3Z4-JGg')
        if price != None:
            prices.append(price.text.replace('\xa0', ' '))
        else:
            prices.append(' ')
    return prices


if __name__ == '__main__':
    link_arr = scrap_services_menu('https://profi.ru/rph/').keys()
    for link in link_arr:
        parsing_info(link)